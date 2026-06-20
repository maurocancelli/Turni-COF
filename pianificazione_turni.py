import streamlit as st
import pandas as pd
import datetime
import io
import gspread
from google.oauth2.service_account import Credentials
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.enums import TA_CENTER
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Pianificazione Turni - COF", layout="wide")

# ─────────────────────────────────────────────
# ACCESSO CON PASSWORD
# ─────────────────────────────────────────────
PASSWORD_ACCESSO = "maddafakka"

FOOTER_HTML = """
    <style>
        .footer-credito {
            position: fixed;
            bottom: 4px;
            left: 0;
            width: 100%;
            text-align: center;
            font-size: 0.7rem;
            color: #999;
            z-index: 100;
            pointer-events: none;
        }
    </style>
    <div class="footer-credito">ideato e realizzato da Mauro Cancelli</div>
"""

if "autenticato" not in st.session_state:
    # Controlla se l'URL contiene il token di sessione (sopravvive ai refresh)
    st.session_state.autenticato = st.query_params.get("auth") == "ok"

if not st.session_state.autenticato:
    st.markdown("### 🔒 Accesso riservato")
    pwd = st.text_input("Inserisci la password:", type="password")
    if st.button("Accedi", type="primary"):
        if pwd == PASSWORD_ACCESSO:
            st.session_state.autenticato = True
            st.query_params["auth"] = "ok"
            st.rerun()
        else:
            st.error("❌ Password errata.")
    st.markdown("---")
    st.caption("ideato e realizzato da Mauro Cancelli")
    st.stop()

st.markdown(FOOTER_HTML, unsafe_allow_html=True)

st.markdown("""
    <style>
        .block-container {
            padding-top: 2.5rem;
            padding-bottom: 1rem;
        }
    </style>
""", unsafe_allow_html=True)
col_titolo, col_codice = st.columns([4, 1])
with col_titolo:
    st.markdown("### Pianificazione Trimestrale Reparto E-commerce")
with col_codice:
    st.markdown("""
        <div style="text-align: right; padding-top: 1.2rem; font-weight: bold; color: #555;">
            COFW-0340
        </div>
    """, unsafe_allow_html=True)

# ─────────────────────────────────────────────
# COSTANTI
# ─────────────────────────────────────────────
GIORNI_LABELS = ["Dom", "Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom"]
GIORNI_CHIAVI = ["Dom_P", "Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom_S"]
GIORNI_BASE   = ["Lunedì", "Martedì", "Mercoledì", "Giovedì", "Venerdì", "Sabato"]
OFFSETS       = [-1, 0, 1, 2, 3, 4, 5, 6]

OPZIONI_TURNO  = ["06:00-13:00", "06:00-13:00*", "07:00-14:00", "07:00-14:00*", "12:30-19:30", "13:00-20:00", "RIPOSO", "MALATTIA", "FERIE", "PERMESSO"]
TARGET_DEFAULT = {"Dom_P": 45, "Lun": 90, "Mar": 75, "Mer": 75, "Gio": 75, "Ven": 90, "Sab": 90, "Dom_S": 45}
TARGET_DOM     = 10

# Domenica: solo turno mattino
TURNO_DOMENICA = "06:00-13:00"

MATRICE_TURNI = {
    0: {1: "06:00-13:00", 2: "06:00-13:00", 3: "13:00-20:00",  4: "12:30-19:30"},
    1: {1: "12:30-19:30", 2: "13:00-20:00", 3: "06:00-13:00",  4: "06:00-13:00"},
    2: {1: "06:00-13:00", 2: "06:00-13:00", 3: "12:30-19:30",  4: "13:00-20:00"},
    3: {1: "13:00-20:00", 2: "12:30-19:30", 3: "06:00-13:00",  4: "06:00-13:00"},
}

ASSENTE = {"RIPOSO", "MALATTIA", "FERIE", "PERMESSO"}

# ─────────────────────────────────────────────
# TIPO ORARIO (visualizzazione contrattuale ridotta)
# I valori "grezzi" salvati nelle settimane restano SEMPRE quelli standard
# (06:00-13:00, 12:30-19:30, 13:00-20:00, 07:00-14:00, con/senza asterisco).
# La traduzione in orario ridotto avviene SOLO in visualizzazione/export
# (Vista Colorata, PDF, Excel), in base al "Tipo Orario" del dipendente.
# ─────────────────────────────────────────────
TIPO_ORARIO_OPZIONI = ["Disponibile", "Contratto 6,15", "Contratto 6,40"]

# mappa: (tipo_orario, turno_grezzo_senza_asterisco) -> orario_visualizzato_senza_asterisco
TRADUZIONE_ORARI = {
    "Disponibile": {
        "06:00-13:00": "06:00-13:00",
        "07:00-14:00": "07:00-14:00",
        "12:30-19:30": "12:30-19:30",
        "13:00-20:00": "13:00-20:00",
    },
    "Contratto 6,15": {
        "06:00-13:00": "06:00-12:15",
        "07:00-14:00": "07:45-14:00",
        "12:30-19:30": "12:30-18:45",
        "13:00-20:00": "13:45-20:00",
    },
    "Contratto 6,40": {
        "06:00-13:00": "06:00-12:40",
        "07:00-14:00": "07:20-14:00",
        "12:30-19:30": "12:30-19:10",
        "13:00-20:00": "13:20-20:00",
    },
}

def traduci_orario_visualizzato(val, tipo_orario):
    """
    Converte un valore turno 'grezzo' (es. '06:00-13:00*') nell'orario da
    mostrare in base al Tipo Orario del dipendente (Standard/Anziano/Recente).
    Le assenze (RIPOSO/FERIE/MALATTIA/PERMESSO) passano invariate.
    """
    if val in ASSENTE:
        return val
    tipo = tipo_orario if tipo_orario in TRADUZIONE_ORARI else "Disponibile"
    asterisco = val.endswith("*")
    base = val[:-1] if asterisco else val
    tradotto = TRADUZIONE_ORARI[tipo].get(base, base)
    return tradotto + "*" if asterisco else tradotto

def ore_turno(val, tipo_orario):
    """
    Restituisce la durata in ore (float) di un turno 'grezzo' per il Tipo
    Orario del dipendente. Le assenze restituiscono 0.
    """
    if val in ASSENTE:
        return 0.0
    orario_reale = traduci_orario_visualizzato(val, tipo_orario)
    base = orario_reale[:-1] if orario_reale.endswith("*") else orario_reale
    try:
        inizio, fine = base.split("-")
        h_in, m_in = (int(x) for x in inizio.split(":"))
        h_fi, m_fi = (int(x) for x in fine.split(":"))
    except Exception:
        return 0.0
    minuti = (h_fi * 60 + m_fi) - (h_in * 60 + m_in)
    return max(minuti, 0) / 60.0

# ─────────────────────────────────────────────
# UTILITY
# ─────────────────────────────────────────────
def pulisci_riposi(val):
    if pd.isna(val) or val is None or str(val).strip() == "":
        return "Nessuno"
    return str(val).strip()

def turno_infrasettimanale(squadra, numero_settimana):
    ciclo = numero_settimana % 4
    return MATRICE_TURNI[ciclo][int(squadra)]

def colora_celle(valore):
    v = str(valore)
    if v == "MALATTIA":  return "background-color:#ffcccc;color:#cc0000;font-weight:bold;"
    if v == "FERIE":     return "background-color:#ffe6cc;color:#cc6600;font-weight:bold;"
    if v == "PERMESSO":  return "background-color:#e6f2ff;color:#0066cc;"
    if v == "RIPOSO":    return "background-color:#f2f2f2;color:#7f7f7f;"
    if "06:00" in v:     return "background-color:#e6ffed;color:#1a7f37;"
    if "12:30" in v or "13:00" in v: return "background-color:#fbefff;color:#8250df;"
    return ""

def genera_pdf_settimana(df, week_num, lun_w, col_labels, definitiva):
    """
    Genera un PDF: una riga per dipendente, due colonne per ogni giorno
    Lun-Dom: colonna sinistra = turno MATTINO (6-13), colonna destra =
    turno POMERIGGIO (12.30-19.30 o 13-20). Solo una delle due e' valorizzata.
    Le assenze (RIPOSO/FERIE/MALATTIA/PERMESSO) sono scritte centrate
    su entrambe le colonne del giorno.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=4*mm, rightMargin=4*mm, topMargin=6*mm, bottomMargin=6*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitoloWeek", parent=styles["Heading1"],
        fontSize=22, textColor=colors.HexColor("#2E7D32"), spaceAfter=2
    )
    status_style_def = ParagraphStyle(
        "StatusDefinitivo", parent=styles["Normal"],
        fontSize=13, textColor=colors.HexColor("#2E7D32"), fontName="Helvetica-Bold", spaceAfter=6
    )
    status_style_prov = ParagraphStyle(
        "StatusProvvisorio", parent=styles["Normal"],
        fontSize=13, textColor=colors.HexColor("#CC6600"), fontName="Helvetica-Bold", spaceAfter=6
    )

    giorni_pdf = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom_S"]
    nomi_giorni_pdf = ["LUNEDI", "MARTEDI", "MERCOLEDI", "GIOVEDI", "VENERDI", "SABATO", "DOMENICA"]

    NOMI_MESI = ["", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                 "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

    dom_p_data = lun_w - datetime.timedelta(days=1)
    dom_s_data = lun_w + datetime.timedelta(days=6)
    if dom_p_data.month == dom_s_data.month:
        periodo = f"dal {dom_p_data.day} al {dom_s_data.day} {NOMI_MESI[dom_s_data.month]}"
    else:
        periodo = (f"dal {dom_p_data.day} {NOMI_MESI[dom_p_data.month]} "
                   f"al {dom_s_data.day} {NOMI_MESI[dom_s_data.month]}")

    elementi = []
    elementi.append(Paragraph(f"WEEK {week_num} &nbsp;&nbsp; {periodo}", title_style))
    if definitiva:
        elementi.append(Paragraph("DEFINITIVO", status_style_def))
    else:
        elementi.append(Paragraph("PROVVISORIO", status_style_prov))
    elementi.append(Spacer(1, 2*mm))

    def fmt_orario(val):
        """Converte un orario (standard o tradotto per Tipo Orario) in
        (testo_breve, fascia). Fascia 'mattino' se inizia prima delle 12."""
        asterisco = val.endswith("*")
        base = val[:-1] if asterisco else val
        try:
            inizio, fine = base.split("-")
            h_in, m_in = inizio.split(":")
            h_fi, m_fi = fine.split(":")
        except Exception:
            return None, None

        def fmt_ora(h, m):
            h = str(int(h))
            return h if m == "00" else f"{h}.{m}"

        txt = f"{fmt_ora(h_in, m_in)}-{fmt_ora(h_fi, m_fi)}"
        if asterisco:
            txt += "*"
        fascia = "mattino" if int(h_in) < 12 else "pomeriggio"
        return txt, fascia

    # ── Header: nome giorno + data ──
    header1 = ["DIPENDENTE"]
    for chiave, nome_g in zip(giorni_pdf, nomi_giorni_pdf):
        lbl = col_labels.get(chiave, nome_g)
        try:
            giorno_num = lbl.split(" ")[1].split("/")[0]
        except Exception:
            giorno_num = ""
        header1.append(f"{nome_g} {giorno_num}")
        header1.append("")

    data_table = [header1]

    # Traccia per ogni cella: tipo di contenuto, per colorare dopo
    cell_kind = {}  # (row_idx, col_idx) -> "assente"/"mattino"/"pomeriggio"/"valore_assenza"

    anagrafica_idx = st.session_state.df_anagrafica.set_index("Nome")

    for r_idx, (_, row) in enumerate(df.iterrows(), start=1):
        riga = [row["Dipendente"]]
        try:
            tipo_orario_dip = anagrafica_idx.at[row["Dipendente"], "Tipo Orario"]
        except KeyError:
            tipo_orario_dip = "Disponibile"
        for gi, chiave in enumerate(giorni_pdf):
            c1 = 1 + gi * 2
            c2 = c1 + 1
            val = traduci_orario_visualizzato(str(row[chiave]), tipo_orario_dip)
            if val in ASSENTE:
                riga.append(val)
                riga.append("")
                cell_kind[(r_idx, gi)] = ("assente", val)
            else:
                txt, fascia = fmt_orario(val)
                if fascia == "mattino":
                    riga.append(txt)
                    riga.append("")
                    cell_kind[(r_idx, gi)] = ("mattino", txt)
                elif fascia == "pomeriggio":
                    riga.append("")
                    riga.append(txt)
                    cell_kind[(r_idx, gi)] = ("pomeriggio", txt)
                else:
                    riga.append(val)
                    riga.append("")
        data_table.append(riga)

    n_cols = len(header1)
    # Colonna mattino piu' larga per ospitare orari contrattuali ridotti (es. 6-12.40)
    col_widths = [49*mm]
    for _ in giorni_pdf:
        col_widths += [16*mm, 18*mm]

    tbl = Table(data_table, colWidths=col_widths, repeatRows=1)

    # ── Stile base: tutto bianco/nero, nessuno sfondo scuro generale ──
    style_cmds = [
        ("FONTSIZE", (0, 0), (-1, -1), 11),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("BACKGROUND", (0, 0), (-1, -1), colors.white),
        ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
        ("BOX", (0, 0), (-1, -1), 1.2, colors.black),
        ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (0, -1), "LEFT"),
        ("LEFTPADDING", (0, 0), (0, -1), 5),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]

    # Merge header per ogni giorno (2 colonne)
    style_cmds.append(("SPAN", (0, 0), (0, 0)))
    for gi in range(len(giorni_pdf)):
        c1 = 1 + gi * 2
        c2 = c1 + 1
        style_cmds.append(("SPAN", (c1, 0), (c2, 0)))
        style_cmds.append(("LINEAFTER", (c2, 0), (c2, -1), 1.2, colors.black))

    # Colori per assenze e turni
    palette_assenza = {
        "RIPOSO":   (colors.HexColor("#F2F2F2"), colors.HexColor("#7F7F7F")),
        "FERIE":    (colors.HexColor("#FFE6CC"), colors.HexColor("#CC6600")),
        "MALATTIA": (colors.HexColor("#FFCCCC"), colors.HexColor("#CC0000")),
        "PERMESSO": (colors.HexColor("#E6F2FF"), colors.HexColor("#0066CC")),
    }

    for (r_idx, gi), (kind, val) in cell_kind.items():
        c1 = 1 + gi * 2
        c2 = c1 + 1
        if kind == "assente":
            bg, fg = palette_assenza[val]
            style_cmds.append(("SPAN", (c1, r_idx), (c2, r_idx)))
            style_cmds.append(("BACKGROUND", (c1, r_idx), (c2, r_idx), bg))
            style_cmds.append(("TEXTCOLOR", (c1, r_idx), (c2, r_idx), fg))
            style_cmds.append(("FONTNAME", (c1, r_idx), (c2, r_idx), "Helvetica-Bold"))
            style_cmds.append(("FONTSIZE", (c1, r_idx), (c2, r_idx), 10))
        elif kind == "mattino":
            style_cmds.append(("BACKGROUND", (c1, r_idx), (c1, r_idx), colors.HexColor("#E6FFED")))
            fs = 11 if len(val) <= 4 else (9.5 if len(val) <= 7 else 8)
            style_cmds.append(("FONTSIZE", (c1, r_idx), (c1, r_idx), fs))
        elif kind == "pomeriggio":
            style_cmds.append(("BACKGROUND", (c2, r_idx), (c2, r_idx), colors.HexColor("#FBEFFF")))
            fs = 9.5 if len(val) <= 11 else 8
            style_cmds.append(("FONTSIZE", (c2, r_idx), (c2, r_idx), fs))

    # Righe alterne bianco/grigio chiarissimo SOLO dove non c'e' gia' un colore assenza
    for r_idx in range(1, len(data_table)):
        if r_idx % 2 == 0:
            for gi in range(len(giorni_pdf)):
                if (r_idx, gi) not in cell_kind or cell_kind[(r_idx, gi)][0] in ("mattino", "pomeriggio"):
                    c1 = 1 + gi * 2
                    c2 = c1 + 1
                    if (r_idx, gi) not in cell_kind:
                        style_cmds.append(("BACKGROUND", (c1, r_idx), (c2, r_idx), colors.HexColor("#FAFAFA")))

    tbl.setStyle(TableStyle(style_cmds))
    elementi.append(tbl)

    doc.build(elementi)
    buffer.seek(0)
    return buffer.getvalue()

def genera_pdf_esposizione(df, week_num, lun_w, col_labels, definitiva):
    """
    Versione 'esposizione' del PDF settimanale: stesso layout di
    genera_pdf_settimana, ma su 2 pagine A4 orizzontale, con righe e
    font maggiorati per massima leggibilita' da lontano.
    Pagina 1: Squadre 1+2. Pagina 2: Squadre 3+4.
    L'intestazione (WEEK/periodo/stato) e' ripetuta su entrambe le pagine.
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(A4),
        leftMargin=1*mm, rightMargin=1*mm, topMargin=6*mm, bottomMargin=6*mm
    )

    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        "TitoloWeekExp", parent=styles["Heading1"],
        fontSize=34, textColor=colors.HexColor("#2E7D32"), fontName="Helvetica-Bold", spaceAfter=2
    )

    giorni_pdf = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom_S"]
    nomi_giorni_pdf = ["LUNEDI", "MARTEDI", "MERCOLEDI", "GIOVEDI", "VENERDI", "SABATO", "DOMENICA"]

    NOMI_MESI = ["", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                 "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]

    dom_p_data = lun_w - datetime.timedelta(days=1)
    dom_s_data = lun_w + datetime.timedelta(days=6)
    if dom_p_data.month == dom_s_data.month:
        periodo = f"dal {dom_p_data.day} al {dom_s_data.day} {NOMI_MESI[dom_s_data.month]}"
    else:
        periodo = (f"dal {dom_p_data.day} {NOMI_MESI[dom_p_data.month]} "
                   f"al {dom_s_data.day} {NOMI_MESI[dom_s_data.month]}")

    def fmt_orario(val):
        """Converte un orario (standard o tradotto per Tipo Orario) in
        (testo_breve, fascia). Fascia 'mattino' se inizia prima delle 12."""
        asterisco = val.endswith("*")
        base = val[:-1] if asterisco else val
        try:
            inizio, fine = base.split("-")
            h_in, m_in = inizio.split(":")
            h_fi, m_fi = fine.split(":")
        except Exception:
            return None, None

        def fmt_ora(h, m):
            h = str(int(h))
            return h if m == "00" else f"{h}.{m}"

        txt = f"{fmt_ora(h_in, m_in)}-{fmt_ora(h_fi, m_fi)}"
        if asterisco:
            txt += "*"
        fascia = "mattino" if int(h_in) < 12 else "pomeriggio"
        return txt, fascia

    palette_assenza = {
        "RIPOSO":   (colors.HexColor("#F2F2F2"), colors.HexColor("#7F7F7F")),
        "FERIE":    (colors.HexColor("#FFE6CC"), colors.HexColor("#CC6600")),
        "MALATTIA": (colors.HexColor("#FFCCCC"), colors.HexColor("#CC0000")),
        "PERMESSO": (colors.HexColor("#E6F2FF"), colors.HexColor("#0066CC")),
    }

    # Header colonne (comune alle due pagine)
    header1 = ["DIPENDENTE"]
    for chiave, nome_g in zip(giorni_pdf, nomi_giorni_pdf):
        lbl = col_labels.get(chiave, nome_g)
        try:
            giorno_num = lbl.split(" ")[1].split("/")[0]
        except Exception:
            giorno_num = ""
        header1.append(f"{nome_g} {giorno_num}")
        header1.append("")

    col_widths = [56*mm]
    for _ in giorni_pdf:
        col_widths += [14*mm, 20*mm]

    anagrafica_idx = st.session_state.df_anagrafica.set_index("Nome")

    def costruisci_tabella(df_gruppo):
        data_table = [header1]
        cell_kind = {}
        for r_idx, (_, row) in enumerate(df_gruppo.iterrows(), start=1):
            riga = [row["Dipendente"]]
            try:
                tipo_orario_dip = anagrafica_idx.at[row["Dipendente"], "Tipo Orario"]
            except KeyError:
                tipo_orario_dip = "Disponibile"
            for gi, chiave in enumerate(giorni_pdf):
                val = traduci_orario_visualizzato(str(row[chiave]), tipo_orario_dip)
                if val in ASSENTE:
                    riga.append(val)
                    riga.append("")
                    cell_kind[(r_idx, gi)] = ("assente", val)
                else:
                    txt, fascia = fmt_orario(val)
                    if fascia == "mattino":
                        riga.append(txt)
                        riga.append("")
                        cell_kind[(r_idx, gi)] = ("mattino", txt)
                    elif fascia == "pomeriggio":
                        riga.append("")
                        riga.append(txt)
                        cell_kind[(r_idx, gi)] = ("pomeriggio", txt)
                    else:
                        riga.append(val)
                        riga.append("")
            data_table.append(riga)

        tbl = Table(data_table, colWidths=col_widths, repeatRows=1)

        style_cmds = [
            ("FONTSIZE", (0, 0), (-1, -1), 13),
            ("FONTSIZE", (0, 0), (-1, 0), 11),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("BACKGROUND", (0, 0), (-1, -1), colors.white),
            ("TEXTCOLOR", (0, 0), (-1, -1), colors.black),
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2E7D32")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("ALIGN", (1, 0), (-1, -1), "CENTER"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#999999")),
            ("BOX", (0, 0), (-1, -1), 1.2, colors.black),
            ("FONTNAME", (0, 1), (0, -1), "Helvetica-Bold"),
            ("ALIGN", (0, 0), (0, -1), "LEFT"),
            ("LEFTPADDING", (0, 0), (0, -1), 6),
            ("TOPPADDING", (0, 0), (-1, -1), 9),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 9),
        ]

        style_cmds.append(("SPAN", (0, 0), (0, 0)))
        for gi in range(len(giorni_pdf)):
            c1 = 1 + gi * 2
            c2 = c1 + 1
            style_cmds.append(("SPAN", (c1, 0), (c2, 0)))
            style_cmds.append(("LINEAFTER", (c2, 0), (c2, -1), 1.2, colors.black))

        for (r_idx, gi), (kind, val) in cell_kind.items():
            c1 = 1 + gi * 2
            c2 = c1 + 1
            if kind == "assente":
                bg, fg = palette_assenza[val]
                style_cmds.append(("SPAN", (c1, r_idx), (c2, r_idx)))
                style_cmds.append(("BACKGROUND", (c1, r_idx), (c2, r_idx), bg))
                style_cmds.append(("TEXTCOLOR", (c1, r_idx), (c2, r_idx), fg))
                style_cmds.append(("FONTNAME", (c1, r_idx), (c2, r_idx), "Helvetica-Bold"))
                style_cmds.append(("FONTSIZE", (c1, r_idx), (c2, r_idx), 12))
            elif kind == "mattino":
                style_cmds.append(("BACKGROUND", (c1, r_idx), (c1, r_idx), colors.HexColor("#E6FFED")))
                fs = 13 if len(val) <= 4 else (11 if len(val) <= 7 else 9)
                style_cmds.append(("FONTSIZE", (c1, r_idx), (c1, r_idx), fs))
            elif kind == "pomeriggio":
                style_cmds.append(("BACKGROUND", (c2, r_idx), (c2, r_idx), colors.HexColor("#FBEFFF")))
                fs = 12 if len(val) <= 9 else (10 if len(val) <= 11 else 9)
                style_cmds.append(("FONTSIZE", (c2, r_idx), (c2, r_idx), fs))

        for r_idx in range(1, len(data_table)):
            if r_idx % 2 == 0:
                for gi in range(len(giorni_pdf)):
                    if (r_idx, gi) not in cell_kind:
                        c1 = 1 + gi * 2
                        c2 = c1 + 1
                        style_cmds.append(("BACKGROUND", (c1, r_idx), (c2, r_idx), colors.HexColor("#FAFAFA")))

        tbl.setStyle(TableStyle(style_cmds))
        return tbl

    def aggiungi_intestazione(elementi, sottotitolo):
        elementi.append(Paragraph(
            f"WEEK {week_num} &nbsp;&nbsp; {periodo}",
            title_style
        ))
        if sottotitolo:
            elementi.append(Spacer(1, 3*mm))
            elementi.append(Paragraph(sottotitolo, ParagraphStyle(
                "Sottotitolo", parent=styles["Normal"],
                fontSize=14, textColor=colors.black, fontName="Helvetica-Bold", spaceAfter=4
            )))
        elementi.append(Spacer(1, 2*mm))

    squadra_int = df["Squadra"].astype(int)
    df_p1 = df[squadra_int.isin([1, 2])].reset_index(drop=True)
    df_p2 = df[squadra_int.isin([3, 4])].reset_index(drop=True)

    elementi = []
    aggiungi_intestazione(elementi, "Squadre 1 e 2")
    elementi.append(costruisci_tabella(df_p1))
    elementi.append(PageBreak())
    aggiungi_intestazione(elementi, "Squadre 3 e 4")
    elementi.append(costruisci_tabella(df_p2))

    doc.build(elementi)
    buffer.seek(0)
    return buffer.getvalue()

def genera_excel_settimana(df, week_num, lun_w, col_labels, definitiva):
    """
    Genera un file Excel (.xlsx): una riga per dipendente, 4 colonne per ogni
    giorno Lun-Dom: [In1, Out1, In2, Out2].
      - Turno 06:00-13:00 o 06:00-13:00* o 07:00-14:00 -> In1/Out1 valorizzate
        (es. "6","13" oppure "7","14"), In2/Out2 vuote.
      - Turno 12:30-19:30 o 13:00-20:00 -> In1/Out1 vuote, In2/Out2 valorizzate
        (es. "12,30","19,30" oppure "13","20").
      - Assenze (RIPOSO/FERIE/MALATTIA/PERMESSO) -> tutte 4 le colonne vuote.
    Header su due righe: riga 1 = nome giorno (merged su 4 colonne),
    riga 2 = vuota.
    """
    giorni_excel = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom_S"]
    nomi_giorni_excel = ["LUNEDI", "MARTEDI", "MERCOLEDI", "GIOVEDI", "VENERDI", "SABATO", "DOMENICA"]

    def split_orario(val):
        """Restituisce (in1, out1, in2, out2) come stringhe, vuote se non applicabile.
        Colonne mattino (in1/out1) se l'orario inizia prima delle 12,
        colonne pomeriggio (in2/out2) altrimenti."""
        base = val[:-1] if val.endswith("*") else val
        try:
            inizio, fine = base.split("-")
            h_in, m_in = inizio.split(":")
            h_fi, m_fi = fine.split(":")
        except Exception:
            return ("", "", "", "")

        def fmt_ora(h, m):
            h = str(int(h))
            return h if m == "00" else f"{h},{m}"

        txt_in, txt_fi = fmt_ora(h_in, m_in), fmt_ora(h_fi, m_fi)
        if int(h_in) < 12:
            return (txt_in, txt_fi, "", "")
        return ("", "", txt_in, txt_fi)

    wb = Workbook()
    ws = wb.active
    ws.title = f"Week {week_num}"

    # ── Stili ──
    header_fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    name_font = Font(bold=True, size=11)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(border_style="thin", color="999999")
    thick = Side(border_style="medium", color="000000")
    border_normal = Border(left=thin, right=thin, top=thin, bottom=thin)

    palette_assenza = {
        "RIPOSO":   PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid"),
        "FERIE":    PatternFill(start_color="FFE6CC", end_color="FFE6CC", fill_type="solid"),
        "MALATTIA": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "PERMESSO": PatternFill(start_color="E6F2FF", end_color="E6F2FF", fill_type="solid"),
    }
    palette_font = {
        "RIPOSO":   Font(color="7F7F7F", bold=True),
        "FERIE":    Font(color="CC6600", bold=True),
        "MALATTIA": Font(color="CC0000", bold=True),
        "PERMESSO": Font(color="0066CC", bold=True),
    }
    fill_mattino = PatternFill(start_color="E6FFED", end_color="E6FFED", fill_type="solid")
    fill_pomeriggio = PatternFill(start_color="FBEFFF", end_color="FBEFFF", fill_type="solid")

    # ── Titolo (riga 1, sopra l'header) ──
    dom_p_data = lun_w - datetime.timedelta(days=1)
    dom_s_data = lun_w + datetime.timedelta(days=6)
    NOMI_MESI = ["", "gennaio", "febbraio", "marzo", "aprile", "maggio", "giugno",
                 "luglio", "agosto", "settembre", "ottobre", "novembre", "dicembre"]
    if dom_p_data.month == dom_s_data.month:
        periodo = f"dal {dom_p_data.day} al {dom_s_data.day} {NOMI_MESI[dom_s_data.month]}"
    else:
        periodo = (f"dal {dom_p_data.day} {NOMI_MESI[dom_p_data.month]} "
                   f"al {dom_s_data.day} {NOMI_MESI[dom_s_data.month]}")
    stato_label = "DEFINITIVO" if definitiva else "PROVVISORIO"

    n_giorni = len(giorni_excel)
    n_cols_tot = 1 + n_giorni * 4  # 1 colonna nomi + 4 per giorno

    ws.cell(row=1, column=1, value=f"WEEK {week_num}   {periodo}   -   {stato_label}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=13, color="2E7D32")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols_tot)

    # ── Header riga 2: nome dipendente + nomi giorni (merge su 4 colonne) ──
    HEADER_ROW = 2
    SUBHEADER_ROW = 3
    DATA_START_ROW = 4

    ws.cell(row=HEADER_ROW, column=1, value="DIPENDENTE")
    ws.cell(row=HEADER_ROW, column=1).fill = header_fill
    ws.cell(row=HEADER_ROW, column=1).font = header_font
    ws.cell(row=HEADER_ROW, column=1).alignment = center
    ws.merge_cells(start_row=HEADER_ROW, start_column=1, end_row=SUBHEADER_ROW, end_column=1)

    for gi, (chiave, nome_g) in enumerate(zip(giorni_excel, nomi_giorni_excel)):
        c1 = 2 + gi * 4
        lbl = col_labels.get(chiave, nome_g)
        try:
            giorno_num = lbl.split(" ")[1].split("/")[0]
        except Exception:
            giorno_num = ""
        cell = ws.cell(row=HEADER_ROW, column=c1, value=f"{nome_g} {giorno_num}")
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        ws.merge_cells(start_row=HEADER_ROW, start_column=c1, end_row=HEADER_ROW, end_column=c1 + 3)
        # Riga 3 (sub-header) vuota ma colorata
        for cc in range(c1, c1 + 4):
            sub = ws.cell(row=SUBHEADER_ROW, column=cc, value="")
            sub.fill = header_fill

    anagrafica_idx = st.session_state.df_anagrafica.set_index("Nome")

    # ── Righe dati ──
    for r_offset, (_, row) in enumerate(df.iterrows()):
        r = DATA_START_ROW + r_offset
        cell_nome = ws.cell(row=r, column=1, value=row["Dipendente"])
        cell_nome.font = name_font
        cell_nome.border = border_normal
        cell_nome.alignment = Alignment(horizontal="left", vertical="center")

        try:
            tipo_orario_dip = anagrafica_idx.at[row["Dipendente"], "Tipo Orario"]
        except KeyError:
            tipo_orario_dip = "Disponibile"

        for gi, chiave in enumerate(giorni_excel):
            c1 = 2 + gi * 4
            val = traduci_orario_visualizzato(str(row[chiave]), tipo_orario_dip)

            if val in ASSENTE:
                for cc in range(c1, c1 + 4):
                    cell = ws.cell(row=r, column=cc, value="")
                    cell.alignment = center
                    cell.fill = palette_assenza[val]
                    cell.font = palette_font[val]
                    cell.border = border_normal
            else:
                in1, out1, in2, out2 = split_orario(val)
                for offset_c, v in enumerate([in1, out1, in2, out2]):
                    cell = ws.cell(row=r, column=c1 + offset_c, value=v)
                    cell.alignment = center
                    cell.border = border_normal
                    if in1 or out1:
                        if offset_c in (0, 1):
                            cell.fill = fill_mattino
                    if in2 or out2:
                        if offset_c in (2, 3):
                            cell.fill = fill_pomeriggio

    # ── Larghezze colonne ──
    ws.column_dimensions["A"].width = 24
    for gi in range(n_giorni):
        for offset_c in range(4):
            col_letter = get_column_letter(2 + gi * 4 + offset_c)
            ws.column_dimensions[col_letter].width = 7

    # ── Bordo scuro tra giorni (separatore visivo) ──
    for gi in range(n_giorni):
        c_last = 2 + gi * 4 + 3  # ultima colonna del giorno
        for r in range(HEADER_ROW, DATA_START_ROW + len(df)):
            cell = ws.cell(row=r, column=c_last)
            existing = cell.border
            cell.border = Border(
                left=existing.left, top=existing.top, bottom=existing.bottom,
                right=thick
            )

    ws.freeze_panes = ws.cell(row=DATA_START_ROW, column=2)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer.getvalue()

# ─────────────────────────────────────────────
# BACKEND PERSISTENZA: GOOGLE SHEETS
# ─────────────────────────────────────────────
SHEET_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

WS_ANAGRAFICA = "anagrafica"
WS_TURNI = "turni"
WS_MODIFICHE = "modifiche"

COLONNE_TURNI = tuple(["Anno", "Week", "Definitiva", "Dipendente", "Contratto", "Squadra"] + GIORNI_CHIAVI)
COLONNE_MODIFICHE = ("Anno", "Week", "Dipendente", "Colonna", "Valore")
COLONNE_ANAGRAFICA = ("Nome", "Contratto", "Squadra", "Riposo 1", "Riposo 2",
                       "Malattia Fino Al", "Ferie W1", "Ferie W2", "Ferie W3", "Tipo Orario")

@st.cache_resource(show_spinner=False)
def get_spreadsheet():
    creds = Credentials.from_service_account_info(
        dict(st.secrets["gcp_service_account"]), scopes=SHEET_SCOPES
    )
    gc = gspread.authorize(creds)
    return gc.open_by_key(st.secrets["sheet"]["id"])

def get_worksheet(nome, colonne):
    """Ritorna il worksheet con quel nome, creandolo (con header) se non esiste."""
    sh = get_spreadsheet()
    try:
        ws = sh.worksheet(nome)
    except gspread.exceptions.WorksheetNotFound:
        ws = sh.add_worksheet(title=nome, rows=200, cols=max(len(colonne), 10))
        ws.update([list(colonne)], "A1")
    return ws

@st.cache_data(ttl=20, show_spinner=False)
def _leggi_worksheet_df(nome, colonne_tuple, _cache_bust=0):
    """Legge un intero worksheet come DataFrame di stringhe."""
    colonne = list(colonne_tuple)
    ws = get_worksheet(nome, colonne)
    valori = ws.get_all_values()
    if not valori or len(valori) < 1:
        return pd.DataFrame(columns=colonne)
    header = valori[0]
    righe = valori[1:]
    df = pd.DataFrame(righe, columns=header)
    for c in colonne:
        if c not in df.columns:
            df[c] = ""
    return df[colonne]

def _scrivi_worksheet_df(nome, colonne, df):
    """Sovrascrive interamente un worksheet con il contenuto del DataFrame."""
    ws = get_worksheet(nome, colonne)
    df = df.copy()
    for c in colonne:
        if c not in df.columns:
            df[c] = ""
    df = df[list(colonne)].fillna("").astype(str)
    valori = [list(colonne)] + df.values.tolist()
    ws.clear()
    ws.update(valori, "A1")
    # Invalida la cache di lettura per questo worksheet
    st.session_state["_cache_bust"] = st.session_state.get("_cache_bust", 0) + 1
    _leggi_worksheet_df.clear()

def _cache_bust():
    return st.session_state.get("_cache_bust", 0)

def rimuovi_settimana_salvata(anno, week):
    """Rimuove la riga salvata per (anno, week) dal worksheet turni."""
    tutto = _leggi_worksheet_df(WS_TURNI, COLONNE_TURNI, _cache_bust())
    if tutto.empty:
        return
    tutto = tutto[~((tutto["Anno"] == str(anno)) & (tutto["Week"] == str(week)))]
    _scrivi_worksheet_df(WS_TURNI, COLONNE_TURNI, tutto)

def is_definitiva(anno, week):
    df = _leggi_worksheet_df(WS_TURNI, COLONNE_TURNI, _cache_bust())
    if df.empty:
        return False
    sel = df[(df["Anno"] == str(anno)) & (df["Week"] == str(week))]
    if sel.empty:
        return False
    return bool(sel.iloc[0]["Definitiva"] == "True")

def ha_modifiche_manuali(anno, week):
    df = _leggi_worksheet_df(WS_MODIFICHE, COLONNE_MODIFICHE, _cache_bust())
    if df.empty:
        return False
    sel = df[(df["Anno"] == str(anno)) & (df["Week"] == str(week))]
    return len(sel) > 0

def salva_settimana(df, anno, week, definitiva):
    tutto = _leggi_worksheet_df(WS_TURNI, COLONNE_TURNI, _cache_bust())
    if not tutto.empty:
        tutto = tutto[~((tutto["Anno"] == str(anno)) & (tutto["Week"] == str(week)))]
    nuove = df.copy()
    nuove["Anno"] = str(anno)
    nuove["Week"] = str(week)
    nuove["Definitiva"] = str(bool(definitiva))
    tutto = pd.concat([tutto, nuove[list(COLONNE_TURNI)]], ignore_index=True)
    _scrivi_worksheet_df(WS_TURNI, COLONNE_TURNI, tutto)

def carica_settimana(anno, week):
    df = _leggi_worksheet_df(WS_TURNI, COLONNE_TURNI, _cache_bust())
    if df.empty:
        return None
    sel = df[(df["Anno"] == str(anno)) & (df["Week"] == str(week))].copy()
    if sel.empty:
        return None
    sel = sel.drop(columns=["Anno", "Week", "Definitiva"])
    sel["Squadra"] = pd.to_numeric(sel["Squadra"], errors="coerce")
    return sel.reset_index(drop=True)

def carica_modifiche(anno, week):
    """Restituisce dict {(nome, colonna): valore} delle modifiche manuali salvate."""
    df = _leggi_worksheet_df(WS_MODIFICHE, COLONNE_MODIFICHE, _cache_bust())
    if df.empty:
        return {}
    sel = df[(df["Anno"] == str(anno)) & (df["Week"] == str(week))]
    return {
        (row["Dipendente"], row["Colonna"]): row["Valore"]
        for _, row in sel.iterrows()
    }

def salva_modifiche(modifiche_dict, anno, week):
    """modifiche_dict: {(nome, colonna): valore}"""
    tutto = _leggi_worksheet_df(WS_MODIFICHE, COLONNE_MODIFICHE, _cache_bust())
    if not tutto.empty:
        tutto = tutto[~((tutto["Anno"] == str(anno)) & (tutto["Week"] == str(week)))]
    if modifiche_dict:
        nuove = pd.DataFrame([
            {"Anno": str(anno), "Week": str(week), "Dipendente": n, "Colonna": c, "Valore": v}
            for (n, c), v in modifiche_dict.items()
        ])
        tutto = pd.concat([tutto, nuove[list(COLONNE_MODIFICHE)]], ignore_index=True)
    _scrivi_worksheet_df(WS_MODIFICHE, COLONNE_MODIFICHE, tutto)

def calcola_modifiche(df_originale, df_attuale, colonne_assenza_only=None):
    """
    Confronta df_attuale con df_originale (generato dall'algoritmo) e
    restituisce dict {(nome, colonna): valore} per le celle diverse.
    Se colonne_assenza_only è True, considera solo modifiche che impostano
    valori in ASSENTE (MALATTIA/FERIE/PERMESSO) o che rimuovono tali valori.
    """
    modifiche = {}
    cols = [c for c in GIORNI_CHIAVI]
    orig_idx = df_originale.set_index("Dipendente")
    att_idx = df_attuale.set_index("Dipendente")
    for nome in att_idx.index:
        if nome not in orig_idx.index:
            continue
        for col in cols:
            v_att = att_idx.at[nome, col]
            v_orig = orig_idx.at[nome, col]
            if str(v_att) != str(v_orig):
                if colonne_assenza_only:
                    if str(v_att) in ASSENTE or str(v_orig) in ASSENTE:
                        modifiche[(nome, col)] = v_att
                else:
                    modifiche[(nome, col)] = v_att
    return modifiche

def applica_modifiche(df, modifiche_dict):
    """Applica le modifiche manuali sopra il df generato dall'algoritmo."""
    if not modifiche_dict:
        return df
    df = df.copy()
    df_idx = df.set_index("Dipendente")
    for (nome, col), val in modifiche_dict.items():
        if nome in df_idx.index and col in df_idx.columns:
            df_idx.at[nome, col] = val
    return df_idx.reset_index()

def verifica_riposo_settimanale(df):
    """
    Controlla che ogni dipendente abbia almeno un giorno di riposo nella
    settimana Dom_P->Sab (7 giorni). Dom_S appartiene alla settimana
    successiva e non viene considerato in questo controllo.
    Restituisce la lista dei nomi che lavorano tutti i 7 giorni.
    """
    giorni_settimana = ["Dom_P", "Lun", "Mar", "Mer", "Gio", "Ven", "Sab"]
    senza_riposo = []
    for _, row in df.iterrows():
        giorni_assenza = sum(1 for chiave in giorni_settimana if str(row[chiave]) in ASSENTE)
        if giorni_assenza == 0:
            senza_riposo.append(row["Dipendente"])
    return senza_riposo

def parse_data_malattia(val):
    if val is None:
        return None
    try:
        parsed = pd.to_datetime(val, dayfirst=True)
        return None if pd.isnull(parsed) else parsed.date()
    except Exception:
        return None


# ─────────────────────────────────────────────
# INIT ANAGRAFICA
# ─────────────────────────────────────────────
def init_anagrafica():
    df = _leggi_worksheet_df(WS_ANAGRAFICA, COLONNE_ANAGRAFICA, _cache_bust())
    if not df.empty:
        df = df.replace("", None)
        for col in ["Riposo 1", "Riposo 2"]:
            df[col] = df[col].apply(pulisci_riposi)
        for col in ["Squadra", "Ferie W1", "Ferie W2", "Ferie W3"]:
            df[col] = pd.to_numeric(df[col], errors="coerce")
            df[col] = df[col].astype("Int64")
            if col != "Squadra":
                df[col] = df[col].astype(object).where(df[col].notna(), None)
        df["Tipo Orario"] = df["Tipo Orario"].apply(
            lambda v: v if v in TIPO_ORARIO_OPZIONI else "Disponibile"
        )
        return df.reset_index(drop=True)

    nomi_base = [
        ("MARVIN MENDOZA","FT",1),("MANUEL MENDOZA","FT",2),
        ("ELAINE ALVES BRAGA","FT",3),("CRISTOPHER BULOSAN","FT",4),
        ("CATERINA CAVALLARI","FT",1),("MICHELA GAVAZZENI","FT",2),
        ("FRANCESCO CUARESMA","FT",3),("BARBARA PERALTA","FT",4),
        ("SIMONA RICCARDI","FT",1),("MARIA MABANTA","FT",2),
        ("ARVIN PASCUA","FT",3),("CAMILLA TOCCHETTI","FT",4),
        ("DAPHNI SPEERING","FT",1),("EUGENE TANADA","FT",2),
        ("RUSSEL CIPRIANO","FT",3),("CARBUNGCAL FELIX","FT",4),
        ("SARA GHITTI","FT",1),("DANIELA BONO","FT",2),
        ("JONIMAE TAN","FT",3),("ROSARIO PETILUNA","FT",4),
        ("LEA MAGTIBAY","FT",1),("BABYJANE MAGTIBAY","FT",2),
        ("NIKITA DONGHI","FT",3),
    ]
    rows = []
    for nome, contratto, sq in nomi_base:
        rows.append({
            "Nome": nome, "Contratto": contratto, "Squadra": sq,
            "Riposo 1": "Nessuno", "Riposo 2": "Nessuno",
            "Malattia Fino Al": None,
            "Ferie W1": None, "Ferie W2": None, "Ferie W3": None,
            "Tipo Orario": "Disponibile",
        })
    df = pd.DataFrame(rows)
    _scrivi_worksheet_df(WS_ANAGRAFICA, COLONNE_ANAGRAFICA, df)
    return df

if "df_anagrafica" not in st.session_state:
    st.session_state.df_anagrafica = init_anagrafica()

def salva_anagrafica(df):
    st.session_state.df_anagrafica = df.copy()
    _scrivi_worksheet_df(WS_ANAGRAFICA, COLONNE_ANAGRAFICA, df)

# ─────────────────────────────────────────────
# LEGGI DOM_S DA FILE STORICO (settimana prima della finestra)
# ─────────────────────────────────────────────
def leggi_dom_s_precedente(week_target, anno_target):
    """
    Restituisce dict nome -> valore_dom_s (stringa) dalla settimana
    immediatamente prima di week_target. Se non esiste ritorna {}.
    """
    w = week_target - 1
    a = anno_target
    if w < 1:
        w = 52
        a -= 1
    df = carica_settimana(a, w)
    if df is None or "Dom_S" not in df.columns:
        return {}
    return {
        row["Dipendente"]: str(row["Dom_S"])
        for _, row in df.iterrows()
        if row.get("Dipendente")
    }

# ─────────────────────────────────────────────
# GENERAZIONE TABELLONE
# dom_s_prec: dict nome -> valore stringa Dom_S settimana precedente
# ─────────────────────────────────────────────
def genera_tabellone(week_num, anno, lunedi, dom_s_prec, target_pct):
    dati = st.session_state.df_anagrafica.to_dict("records")
    dati = [d for d in dati if d.get("Nome") and str(d.get("Nome")).strip()]
    n_totale = len(dati)

    data_dom_p = lunedi - datetime.timedelta(days=1)
    data_dom_s = lunedi + datetime.timedelta(days=6)
    dip_map = {d["Nome"]: d for d in dati}

    rows = []
    for dip in dati:
        nome = dip["Nome"]
        ferie_set = set()
        for fk in ["Ferie W1", "Ferie W2", "Ferie W3"]:
            v = dip.get(fk)
            if v is not None:
                try:
                    ferie_set.add(int(v))
                except Exception:
                    pass
        in_ferie = week_num in ferie_set
        in_ferie_succ = (week_num + 1) in ferie_set
        data_mal = parse_data_malattia(dip.get("Malattia Fino Al"))
        rows.append({
            "Dipendente": nome, "Contratto": dip["Contratto"], "Squadra": dip["Squadra"],
            "_in_ferie": in_ferie, "_in_ferie_succ": in_ferie_succ, "_data_mal": data_mal,
            "Dom_P": None, "Lun": None, "Mar": None, "Mer": None,
            "Gio": None, "Ven": None, "Sab": None, "Dom_S": None,
        })

    df = pd.DataFrame(rows)

    # ── Turni infrasettimanali Lun-Sab ──
    for idx, row in df.iterrows():
        dip = dip_map[row["Dipendente"]]
        t_base = turno_infrasettimanale(dip["Squadra"], week_num)
        for chiave, offset in zip(GIORNI_CHIAVI[1:7], OFFSETS[1:7]):
            data_g = lunedi + datetime.timedelta(days=offset)
            in_mal = (row["_data_mal"] is not None) and (data_g <= row["_data_mal"])
            if in_mal:
                df.at[idx, chiave] = "MALATTIA"
            elif row["_in_ferie"]:
                df.at[idx, chiave] = "FERIE"
            else:
                df.at[idx, chiave] = t_base
    # ── Dom_P: copia esatta della Dom_S della settimana precedente ──
    # ECCEZIONE: chi è in FERIE questa settimana lavora SEMPRE Dom_P
    # (è l'ultimo giorno prima di partire), a prescindere dalla rotazione.
    for idx, row in df.iterrows():
        data_mal = row["_data_mal"]
        in_mal_dom_p = (data_mal is not None) and (data_dom_p <= data_mal)
        if in_mal_dom_p:
            df.at[idx, "Dom_P"] = "MALATTIA"
        elif row["_in_ferie"]:
            # Lavora obbligatoriamente la domenica prima di partire in ferie
            # Asterisco per segnalare che è un turno "forzato" da non confondere
            df.at[idx, "Dom_P"] = TURNO_DOMENICA + "*"
        else:
            val_prec = dom_s_prec.get(row["Dipendente"], None)
            if val_prec is None:
                # Nessuna info storica: default mattino
                df.at[idx, "Dom_P"] = TURNO_DOMENICA
            else:
                # Copia esatta — stesso valore della Dom_S settimana scorsa
                df.at[idx, "Dom_P"] = val_prec

    # ── Riposi PT (priorità: giorni fissi rispettati sempre) ──
    # Lavora Dom_P → 2 riposi Lun-Sab (entrambi i giorni fissi) = 5 gg lavoro
    # Non lavora Dom_P → 1 solo riposo Lun-Sab (domenica già bruciata) = 5 gg lavoro
    for idx, row in df.iterrows():
        if row["Contratto"] != "PT" or row["_in_ferie"]:
            continue
        dip = dip_map[row["Dipendente"]]
        riposi = [r for r in [dip.get("Riposo 1"), dip.get("Riposo 2")]
                  if r and r != "Nessuno"]
        if not riposi:
            continue

        dom_p_val = str(df.at[idx, "Dom_P"])
        ha_lavorato_dom = dom_p_val not in ASSENTE

        def target_di(g):
            for k, off in zip(GIORNI_CHIAVI[1:7], OFFSETS[1:7]):
                if GIORNI_BASE[off] == g:
                    return target_pct.get(k, 0.75)
            return 0.75

        if len(riposi) == 2:
            t1, t2 = target_di(riposi[0]), target_di(riposi[1])
            # giorno col target minore = meno necessario = candidato a riposo prioritario
            r_prim = riposi[0] if t1 <= t2 else riposi[1]
            r_sec  = riposi[1] if t1 <= t2 else riposi[0]
            # ha lavorato domenica → 2 riposi infrasettimanali (entrambi i giorni fissi)
            # non ha lavorato domenica → 1 solo riposo (domenica già bruciata)
            da_app = [r_prim, r_sec] if ha_lavorato_dom else [r_prim]
        else:
            da_app = riposi

        for g in da_app:
            for chiave, offset in zip(GIORNI_CHIAVI[1:7], OFFSETS[1:7]):
                if GIORNI_BASE[offset] == g:
                    if df.at[idx, chiave] not in {"MALATTIA", "FERIE"}:
                        df.at[idx, chiave] = "RIPOSO"
                    break

    # ── Riposo FT (un giorno Lun-Sab, completa dopo i PT) ──
    # Solo per chi LAVORA Dom_P: ha già il riposo domenicale, quindi
    # gli assegniamo un riposo infrasettimanale.
    # Chi NON lavora Dom_P: il suo riposo è la domenica stessa, nessun
    # altro riposo Lun-Sab.
    # Calcolato DOPO i riposi PT, così il surplus di copertura riflette
    # già i giorni fissi liberati dai part-time.
    for idx, row in df.iterrows():
        if row["Contratto"] != "FT" or row["_in_ferie"]:
            continue
        if all(df.at[idx, k] == "MALATTIA" for k in GIORNI_CHIAVI[1:7]):
            continue
        dom_p_val = str(df.at[idx, "Dom_P"])
        # Riposo infrasettimanale solo se ha lavorato Dom_P
        if dom_p_val in ASSENTE:
            continue
        miglior = None
        max_sur = -9999
        for chiave in GIORNI_CHIAVI[1:7]:
            if df.at[idx, chiave] in ASSENTE:
                continue
            lav = (~df[chiave].isin(ASSENTE)).sum()
            sur = lav - n_totale * target_pct.get(chiave, 0.75)
            if sur > max_sur:
                max_sur = sur
                miglior = chiave
        if miglior:
            df.at[idx, miglior] = "RIPOSO"

    # ── Dom_S: rotazione rispetto a Dom_P della STESSA settimana ──
    for idx, row in df.iterrows():
        data_mal = row["_data_mal"]
        in_mal_dom_s = (data_mal is not None) and (data_dom_s <= data_mal)
        if in_mal_dom_s:
            df.at[idx, "Dom_S"] = "MALATTIA"
        elif row["_in_ferie"]:
            # Chi è in ferie: lavora Dom_P (già assegnata sopra), riposa Dom_S
            df.at[idx, "Dom_S"] = "RIPOSO"
        elif row["_in_ferie_succ"]:
            # Chi parte in ferie la settimana successiva: questa Dom_S è il suo
            # ultimo giorno di lavoro prima di partire (= Dom_P della settimana
            # successiva). Sempre 06:00-13:00* per evidenziarlo, anche in vista.
            df.at[idx, "Dom_S"] = TURNO_DOMENICA + "*"
        else:
            dom_p_val = str(df.at[idx, "Dom_P"])
            if dom_p_val in ASSENTE:
                # Non ha lavorato Dom_P → lavora Dom_S
                df.at[idx, "Dom_S"] = TURNO_DOMENICA
            else:
                # Ha lavorato Dom_P → riposa Dom_S
                df.at[idx, "Dom_S"] = "RIPOSO"

    # ── Garantisci ALMENO TARGET_DOM lavoratori in Dom_S ──
    # Se sono di più (es. per ferie/rotazioni che liberano più persone),
    # va bene così: l'utente può aggiustare manualmente se serve,
    # ed è un margine utile per malattie dell'ultimo minuto.
    lav = (~df["Dom_S"].isin(ASSENTE)).sum()
    mancanti = TARGET_DOM - lav
    if mancanti > 0:
        for idx in df[(df["Dom_S"] == "RIPOSO") & (~df["_in_ferie"])].index:
            if mancanti <= 0:
                break
            df.at[idx, "Dom_S"] = TURNO_DOMENICA
            mancanti -= 1

    df = df.drop(columns=["_in_ferie", "_in_ferie_succ", "_data_mal"])
    return df[["Dipendente", "Contratto", "Squadra"] + GIORNI_CHIAVI]

# ─────────────────────────────────────────────
# RIFERIMENTI TEMPORALI
# ─────────────────────────────────────────────
oggi = datetime.date.today()
lunedi_corrente = oggi - datetime.timedelta(days=oggi.weekday())
iso_c = lunedi_corrente.isocalendar()
week_corrente = iso_c[1]
anno_corrente = iso_c[0]

settimane = []
for delta in range(-1, 14):
    lun = lunedi_corrente + datetime.timedelta(weeks=delta)
    iso = lun.isocalendar()
    settimane.append((iso[0], iso[1], lun))

# ─────────────────────────────────────────────
# PARAMETRI (stato persistente, valori usati dall'algoritmo)
# ─────────────────────────────────────────────
if "target_pct_confermato" not in st.session_state:
    st.session_state.target_pct_confermato = {k: v / 100 for k, v in TARGET_DEFAULT.items()}
if "pieces_ora" not in st.session_state:
    st.session_state.pieces_ora = 100

target_pct = st.session_state.target_pct_confermato
pieces_ora = st.session_state.pieces_ora


# ─────────────────────────────────────────────
# TABS PRINCIPALI
# ─────────────────────────────────────────────
tab_turni, tab_anagrafica, tab_parametri = st.tabs(
    ["📅 Turni Settimanali", "📋 Gestione Anagrafica", "⚙️ Parametri"]
)

# ══════════════════════════════════════════════
# SCHEDA — TURNI SETTIMANALI
# ══════════════════════════════════════════════
with tab_turni:
    if st.session_state.df_anagrafica.empty:
        st.warning("⚠️ Aggiungi prima i dipendenti nell'Anagrafica.")
    else:
        # ── Costruisci labels (con indicatore modifiche manuali) ──
        labels_week = []
        for anno_w, week_w, lun_w in settimane:
            dom_p = lun_w - datetime.timedelta(days=1)
            dom_s = lun_w + datetime.timedelta(days=6)
            flag  = "🔒" if is_definitiva(anno_w, week_w) else "📝"
            mod   = "✏️" if ha_modifiche_manuali(anno_w, week_w) else ""
            cur   = " ◀" if (anno_w == anno_corrente and week_w == week_corrente) else ""
            labels_week.append(f"{flag}{mod} W{week_w} ({dom_p.day}/{dom_p.month}–{dom_s.day}/{dom_s.month}){cur}")

        # ── Pre-genera la catena completa ──
        # Per ogni settimana: genera il tabellone "pulito" (senza modifiche manuali),
        # poi applica sopra le eventuali modifiche manuali salvate.
        # dom_s_map per la catena usa SEMPRE il tabellone con modifiche applicate,
        # cosi' la propagazione tiene conto di quello che l'utente ha effettivamente deciso.
        tabelloni = {}        # (anno, week) -> df CON modifiche applicate (quello mostrato)
        tabelloni_puliti = {} # (anno, week) -> df SENZA modifiche (output puro algoritmo)

        for j, (anno_w, week_w, lun_w) in enumerate(settimane):
            # dom_s_map dalla settimana precedente nella catena (con modifiche)
            if j == 0:
                dom_s_map = leggi_dom_s_precedente(week_w, anno_w)
            else:
                anno_pw, week_pw, _ = settimane[j - 1]
                df_prec = tabelloni.get((anno_pw, week_pw))
                if df_prec is not None and "Dom_S" in df_prec.columns:
                    dom_s_map = {
                        rrow["Dipendente"]: str(rrow["Dom_S"])
                        for _, rrow in df_prec.iterrows()
                        if rrow.get("Dipendente")
                    }
                else:
                    dom_s_map = leggi_dom_s_precedente(week_w, anno_w)

            definitiva_w = is_definitiva(anno_w, week_w)
            df_salvato = carica_settimana(anno_w, week_w)
            modifiche = carica_modifiche(anno_w, week_w)

            if definitiva_w and df_salvato is not None:
                # DEFINITIVA: usa esattamente quanto salvato, nessun ricalcolo.
                # Forza comunque Dom_P coerente con la catena (propagazione domeniche).
                df_pulito = df_salvato.copy()
                for ridx, rrow in df_pulito.iterrows():
                    nome = rrow.get("Dipendente")
                    if nome:
                        val = dom_s_map.get(nome, None)
                        if val is not None:
                            df_pulito.at[ridx, "Dom_P"] = val
                df_con_mod = df_pulito.copy()
            else:
                # PROVVISORIA (salvata o no): rigenera da zero con l'algoritmo,
                # poi applica le modifiche manuali sopra.
                df_pulito = genera_tabellone(week_w, anno_w, lun_w, dom_s_map, target_pct)
                df_con_mod = applica_modifiche(df_pulito, modifiche)

            tabelloni_puliti[(anno_w, week_w)] = df_pulito
            tabelloni[(anno_w, week_w)] = df_con_mod

        # ── Renderizza i tab ──
        tabs_week = st.tabs(labels_week)

        for i, (t_week, (anno_w, week_w, lun_w)) in enumerate(zip(tabs_week, settimane)):
            with t_week:
                col_labels = {}
                rinomina_exp = {}
                for chiave, label, offset in zip(GIORNI_CHIAVI, GIORNI_LABELS, OFFSETS):
                    data_g = lun_w + datetime.timedelta(days=offset)
                    lbl = f"{label} {data_g.day}/{data_g.month}"
                    col_labels[chiave] = lbl
                    rinomina_exp[chiave] = lbl

                definitiva = is_definitiva(anno_w, week_w)
                df_calcolato = tabelloni[(anno_w, week_w)].copy()
                ha_mod = ha_modifiche_manuali(anno_w, week_w)

                if definitiva:
                    config_turni = {
                        "Contratto": None,
                        "Squadra": None,
                    }
                    config_turni.update({
                        chiave: st.column_config.SelectboxColumn(
                            col_labels[chiave],
                            options=OPZIONI_TURNO,
                            disabled=(chiave == "Dom_P")
                        )
                        for chiave in GIORNI_CHIAVI
                    })
                    msg = "🔒 **Settimana DEFINITIVA** — consegnata. Puoi modificare solo MALATTIA/FERIE/PERMESSO; gli orari altrui non cambiano."
                    if ha_mod:
                        msg += "  \n✏️ *Sono presenti modifiche manuali rispetto alla generazione originale.*"
                    st.success(msg)
                else:
                    config_turni = {
                        "Contratto": None,
                        "Squadra": None,
                    }
                    config_turni.update({
                        chiave: st.column_config.SelectboxColumn(
                            col_labels[chiave],
                            options=OPZIONI_TURNO,
                            disabled=(chiave == "Dom_P" and i > 0)
                        )
                        for chiave in GIORNI_CHIAVI
                    })
                    msg = "📝 **Settimana PROVVISORIA** — generata automaticamente, si adatta a malattie/ferie/permessi inseriti."
                    if ha_mod:
                        msg += "  \n✏️ *Sono presenti modifiche manuali, riapplicate ad ogni rigenerazione.*"
                    st.info(msg)

                col_sticker, col_editor = st.columns([1, 30])
                with col_sticker:
                    st.markdown(f"""
                        <div style="
                            writing-mode: vertical-rl;
                            transform: rotate(180deg);
                            background-color: #2E7D32;
                            color: white;
                            font-weight: bold;
                            font-size: 1.8rem;
                            letter-spacing: 2px;
                            border-radius: 6px;
                            padding: 8px 4px;
                            text-align: center;
                            height: 100%;
                            min-height: 200px;
                            display: flex;
                            align-items: center;
                            justify-content: flex-end;
                            padding-bottom: 16px;
                        ">
                            WEEK {week_w}
                        </div>
                    """, unsafe_allow_html=True)
                with col_editor:
                    n_righe = len(df_calcolato)
                    df_modificato = st.data_editor(
                        df_calcolato,
                        column_config=config_turni,
                        width="stretch",
                        hide_index=True,
                        height=(n_righe + 1) * 35 + 3,
                        key=f"editor_{anno_w}_{week_w}"
                    )

                col1, col2, col3, col4, col5 = st.columns(5)

                # ── Controllo riposo settimanale (eseguito ad ogni rerun, mostra alert se necessario) ──
                senza_riposo = verifica_riposo_settimanale(df_modificato)
                if senza_riposo:
                    nomi_str = ", ".join(senza_riposo)
                    st.error(
                        f"⚠️ **Attenzione**: {nomi_str} risulta/risultano programmato/i a lavorare "
                        f"tutti i 7 giorni della settimana (Dom→Sab) senza alcun riposo, ferie, "
                        f"malattia o permesso. Assegna almeno un giorno di riposo prima di salvare."
                    )

                with col1:
                    if not definitiva:
                        if st.button("🔒 Blocca come Definitiva", type="primary",
                                     width="stretch", key=f"blocca_{anno_w}_{week_w}",
                                     disabled=bool(senza_riposo)):
                            salva_settimana(df_modificato, anno_w, week_w, definitiva=True)
                            st.success("✅ Bloccata!")
                            st.rerun()
                    else:
                        if st.button("🔓 Sblocca (torna Provvisoria)", type="secondary",
                                     width="stretch", key=f"sblocca_{anno_w}_{week_w}"):
                            salva_settimana(df_modificato, anno_w, week_w, definitiva=False)
                            st.success("↩️ Sbloccata!")
                            st.rerun()
                with col2:
                    if st.button("💾 Salva Modifiche", width="stretch",
                                 key=f"salva_{anno_w}_{week_w}",
                                 disabled=bool(senza_riposo)):
                        if definitiva:
                            # Salva il file intero così com'è (orari fissi + assenze modificate)
                            salva_settimana(df_modificato, anno_w, week_w, definitiva=True)
                            # Traccia comunque le modifiche di assenza per il badge
                            df_pulito = tabelloni_puliti[(anno_w, week_w)]
                            mod = calcola_modifiche(df_pulito, df_modificato, colonne_assenza_only=True)
                            salva_modifiche(mod, anno_w, week_w)
                        else:
                            # Calcola il diff rispetto al tabellone "pulito" (algoritmo puro)
                            df_pulito = tabelloni_puliti[(anno_w, week_w)]
                            mod = calcola_modifiche(df_pulito, df_modificato, colonne_assenza_only=False)
                            salva_modifiche(mod, anno_w, week_w)
                            # Salva anche il risultato corrente (per coerenza catena domeniche)
                            salva_settimana(df_modificato, anno_w, week_w, definitiva=False)
                        st.success("✅ Salvato!")
                        st.rerun()
                with col3:
                    excel_data = genera_excel_settimana(df_modificato, week_w, lun_w, col_labels, definitiva)
                    st.download_button(
                        label="📊 Esporta Excel",
                        data=excel_data,
                        file_name=f"Turni_W{week_w}_{anno_w}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        width="stretch",
                        key=f"down_{anno_w}_{week_w}"
                    )
                with col4:
                    pdf_data = genera_pdf_settimana(df_modificato, week_w, lun_w, col_labels, definitiva)
                    st.download_button(
                        label="🖨️ Esporta PDF",
                        data=pdf_data,
                        file_name=f"Turni_W{week_w}_{anno_w}.pdf",
                        mime="application/pdf",
                        width="stretch",
                        key=f"pdf_{anno_w}_{week_w}"
                    )
                with col5:
                    pdf_esposizione_data = genera_pdf_esposizione(df_modificato, week_w, lun_w, col_labels, definitiva)
                    st.download_button(
                        label="📌 PDF per esposizione",
                        data=pdf_esposizione_data,
                        file_name=f"Turni_W{week_w}_{anno_w}_esposizione.pdf",
                        mime="application/pdf",
                        width="stretch",
                        key=f"pdf_exp_{anno_w}_{week_w}"
                    )

                if not definitiva and ha_mod:
                    if st.button("🗑️ Rimuovi modifiche manuali (rigenera da zero)",
                                 width="stretch", key=f"reset_mod_{anno_w}_{week_w}"):
                        salva_modifiche({}, anno_w, week_w)
                        rimuovi_settimana_salvata(anno_w, week_w)
                        st.success("♻️ Modifiche manuali rimosse, settimana rigenerata.")
                        st.rerun()

                st.write("**Vista Colorata (formato stampa):**")

                giorni_vista = ["Lun", "Mar", "Mer", "Gio", "Ven", "Sab", "Dom_S"]
                nomi_giorni_vista = ["LUNEDI", "MARTEDI", "MERCOLEDI", "GIOVEDI", "VENERDI", "SABATO", "DOMENICA"]

                def fmt_orario_vista(val):
                    """
                    Converte un valore turno (standard o tradotto per Tipo Orario)
                    in (testo_breve, fascia). Fascia 'mattino' se l'ora di inizio
                    è prima delle 12, 'pomeriggio' altrimenti.
                    """
                    asterisco = val.endswith("*")
                    base = val[:-1] if asterisco else val
                    try:
                        inizio, fine = base.split("-")
                        h_in, m_in = inizio.split(":")
                        h_fi, m_fi = fine.split(":")
                    except Exception:
                        return None, None

                    def fmt_ora(h, m):
                        h = str(int(h))
                        return h if m == "00" else f"{h}.{m}"

                    txt = f"{fmt_ora(h_in, m_in)}-{fmt_ora(h_fi, m_fi)}"
                    if asterisco:
                        txt += "*"
                    fascia = "mattino" if int(h_in) < 12 else "pomeriggio"
                    return txt, fascia

                df_pulito_corrente = tabelloni_puliti[(anno_w, week_w)]
                pulito_idx = df_pulito_corrente.set_index("Dipendente")

                palette_vista = {
                    "RIPOSO":   "background-color:#f2f2f2;color:#7f7f7f;",
                    "FERIE":    "background-color:#ffe6cc;color:#cc6600;",
                    "MALATTIA": "background-color:#ffcccc;color:#cc0000;",
                    "PERMESSO": "background-color:#e6f2ff;color:#0066cc;",
                    "mattino":    "background-color:#e6ffed;color:#1a7f37;",
                    "pomeriggio": "background-color:#fbefff;color:#8250df;",
                }
                EVIDENZIA_GIALLO = "#FFF59D"

                def sostituisci_bg(style, nuovo_bg):
                    if "background-color:" not in style:
                        return f"background-color:{nuovo_bg};" + style
                    vecchio = style.split("background-color:")[1].split(";")[0]
                    return style.replace(vecchio, nuovo_bg)

                # ── Costruzione tabella HTML (celle unite per le assenze) ──
                html = ['<table style="width:100%;border-collapse:collapse;font-size:0.85rem;text-align:center;">']

                # Header: riga con nome giorno + data, ogni giorno occupa 2 colonne (☀️/🌙)
                html.append('<tr>')
                html.append(
                    '<th style="border:1px solid #999;padding:4px 8px;background:#2E7D32;'
                    'color:white;text-align:left;">Dipendente</th>'
                )
                for chiave, nome_g in zip(giorni_vista, nomi_giorni_vista):
                    data_g = lun_w + datetime.timedelta(days=OFFSETS[GIORNI_CHIAVI.index(chiave)])
                    html.append(
                        f'<th colspan="2" style="border:1px solid #999;padding:4px;'
                        f'background:#2E7D32;color:white;border-left:2px solid #000;">'
                        f'{nome_g} {data_g.day}</th>'
                    )
                html.append('</tr>')

                # Sub-header ☀️ / 🌙
                html.append('<tr>')
                html.append('<th style="border:1px solid #999;padding:2px;background:#f5f5f5;"></th>')
                for chiave in giorni_vista:
                    html.append(
                        '<th style="border:1px solid #999;padding:2px;background:#f5f5f5;'
                        'border-left:2px solid #000;">☀️</th>'
                        '<th style="border:1px solid #999;padding:2px;background:#f5f5f5;">🌙</th>'
                    )
                html.append('</tr>')

                anagrafica_idx = st.session_state.df_anagrafica.set_index("Nome")

                # Righe dati
                for r_idx, riga in df_modificato.iterrows():
                    nome_dip = riga["Dipendente"]
                    try:
                        tipo_orario_dip = anagrafica_idx.at[nome_dip, "Tipo Orario"]
                    except KeyError:
                        tipo_orario_dip = "Disponibile"
                    html.append('<tr>')
                    html.append(
                        f'<td style="border:1px solid #999;padding:4px 8px;font-weight:bold;'
                        f'text-align:left;white-space:nowrap;">{nome_dip}</td>'
                    )
                    for chiave in giorni_vista:
                        val = str(riga[chiave])
                        try:
                            val_pulito = str(pulito_idx.at[nome_dip, chiave])
                        except KeyError:
                            val_pulito = val
                        cella_modificata = (not definitiva) and (val != val_pulito)
                        val_vis = traduci_orario_visualizzato(val, tipo_orario_dip)

                        if val_vis in ASSENTE:
                            style = palette_vista[val_vis]
                            if cella_modificata:
                                style = sostituisci_bg(style, EVIDENZIA_GIALLO)
                            html.append(
                                f'<td colspan="2" style="border:1px solid #999;padding:4px;'
                                f'font-weight:bold;border-left:2px solid #000;{style}">{val_vis}</td>'
                            )
                        else:
                            txt, fascia = fmt_orario_vista(val_vis)
                            if fascia == "mattino":
                                style_m = palette_vista["mattino"]
                                if cella_modificata:
                                    style_m = sostituisci_bg(style_m, EVIDENZIA_GIALLO)
                                style_p = ""
                                txt_m, txt_p = txt, ""
                            elif fascia == "pomeriggio":
                                style_p = palette_vista["pomeriggio"]
                                if cella_modificata:
                                    style_p = sostituisci_bg(style_p, EVIDENZIA_GIALLO)
                                style_m = ""
                                txt_m, txt_p = "", txt
                            else:
                                style_m, style_p = "", ""
                                if cella_modificata:
                                    style_m = f"background-color:{EVIDENZIA_GIALLO};"
                                txt_m, txt_p = val_vis, ""

                            html.append(
                                f'<td style="border:1px solid #999;padding:4px;'
                                f'font-weight:bold;border-left:2px solid #000;{style_m}">{txt_m}</td>'
                            )
                            html.append(
                                f'<td style="border:1px solid #999;padding:4px;'
                                f'font-weight:bold;{style_p}">{txt_p}</td>'
                            )
                    html.append('</tr>')

                html.append('</table>')
                st.markdown("".join(html), unsafe_allow_html=True)


                st.write("**Stima Volumi Giornalieri:**")
                anagrafica_idx_report = st.session_state.df_anagrafica.set_index("Nome")
                report = []
                for chiave in GIORNI_CHIAVI:
                    op_m = (df_modificato[chiave].isin(["06:00-13:00", "06:00-13:00*", "07:00-14:00", "07:00-14:00*"])).sum()
                    op_p = (df_modificato[chiave].isin(["12:30-19:30", "13:00-20:00"])).sum()
                    ore_tot = 0.0
                    for _, riga in df_modificato.iterrows():
                        try:
                            tipo_orario_dip = anagrafica_idx_report.at[riga["Dipendente"], "Tipo Orario"]
                        except KeyError:
                            tipo_orario_dip = "Disponibile"
                        ore_tot += ore_turno(str(riga[chiave]), tipo_orario_dip)
                    report.append({
                        "Giorno": col_labels[chiave],
                        "Mattina (06-13)": int(op_m),
                        "Pomeriggio": int(op_p),
                        "Tot. Operatori": int(op_m + op_p),
                        "Totale Ore": round(ore_tot, 2),
                        "Pezzi Stimati": int(ore_tot * pieces_ora),
                    })
                df_report = pd.DataFrame(report).set_index("Giorno").T
                styler = (
                    df_report.style
                    .set_table_styles([
                        {"selector": "th", "props": [("text-align", "right"), ("border", "1px solid #333")]},
                        {"selector": "td", "props": [("text-align", "right"), ("border", "1px solid #333")]},
                        {"selector": "table", "props": [("border-collapse", "collapse"), ("width", "100%")]},
                    ])
                    .set_properties(**{"text-align": "right"})
                )
                st.table(styler)

# ══════════════════════════════════════════════
# SCHEDA 1 — ANAGRAFICA
# ══════════════════════════════════════════════
with tab_anagrafica:
    st.subheader("👥 Lista Personale e Assenze Programmate")
    df_show = st.session_state.df_anagrafica.copy()
    if "Malattia Fino Al" in df_show.columns:
        df_show["Malattia Fino Al"] = pd.to_datetime(df_show["Malattia Fino Al"], errors="coerce", dayfirst=True).dt.date

    config_anagrafica = {
        "Contratto": st.column_config.SelectboxColumn("Contratto", options=["FT", "PT"], required=True),
        "Squadra": st.column_config.NumberColumn("Squadra", min_value=1, max_value=4, step=1, required=True),
        "Riposo 1": st.column_config.SelectboxColumn("Riposo 1 (PT)", options=["Nessuno"] + GIORNI_BASE),
        "Riposo 2": st.column_config.SelectboxColumn("Riposo 2 (PT)", options=["Nessuno"] + GIORNI_BASE),
        "Malattia Fino Al": st.column_config.DateColumn("Malattia Fino Al", format="DD/MM/YYYY"),
        "Ferie W1": st.column_config.NumberColumn("Ferie W1 (N. Sett. ISO)", min_value=1, max_value=53),
        "Ferie W2": st.column_config.NumberColumn("Ferie W2 (N. Sett. ISO)", min_value=1, max_value=53),
        "Ferie W3": st.column_config.NumberColumn("Ferie W3 (N. Sett. ISO)", min_value=1, max_value=53),
        "Tipo Orario": st.column_config.SelectboxColumn("Tipo Orario", options=TIPO_ORARIO_OPZIONI, required=True),
    }
    df_editato = st.data_editor(
        df_show, column_config=config_anagrafica,
        num_rows="dynamic", width="stretch",
        hide_index=True, key="anagrafica_editor"
    )
    if st.button("💾 Salva Modifiche Anagrafica", type="primary", width="stretch"):
        salva_anagrafica(df_editato)
        st.success("✅ Anagrafica aggiornata!")

    st.divider()
    col_add, col_del = st.columns(2)

    with col_add:
        st.subheader("➕ Aggiungi Dipendente")
        with st.container(border=True):
            nuovo_nome      = st.text_input("Nome e Cognome")
            nuovo_contratto = st.selectbox("Contratto", ["FT", "PT"])
            nuova_squadra   = st.selectbox("Squadra", [1, 2, 3, 4])
            nuovo_r1 = st.selectbox("Riposo Fisso 1 (PT)", ["Nessuno"] + GIORNI_BASE)
            nuovo_r2 = st.selectbox("Riposo Fisso 2 (PT)", ["Nessuno"] + GIORNI_BASE)
            nuovo_tipo_orario = st.selectbox("Tipo Orario", TIPO_ORARIO_OPZIONI)
            if st.button("Aggiungi", width="stretch"):
                if not nuovo_nome.strip():
                    st.error("Inserisci un nome valido!")
                else:
                    nuova_riga = {
                        "Nome": nuovo_nome.upper(), "Contratto": nuovo_contratto,
                        "Squadra": nuova_squadra,
                        "Riposo 1": nuovo_r1 if nuovo_contratto == "PT" else "Nessuno",
                        "Riposo 2": nuovo_r2 if nuovo_contratto == "PT" else "Nessuno",
                        "Malattia Fino Al": None,
                        "Ferie W1": None, "Ferie W2": None, "Ferie W3": None,
                        "Tipo Orario": nuovo_tipo_orario,
                    }
                    nuovo_df = pd.concat([st.session_state.df_anagrafica, pd.DataFrame([nuova_riga])], ignore_index=True)
                    salva_anagrafica(nuovo_df)
                    st.success(f"✅ {nuovo_nome.upper()} aggiunto!")
                    st.rerun()

    with col_del:
        st.subheader("🗑️ Rimuovi Dipendente")
        with st.container(border=True):
            lista_nomi = st.session_state.df_anagrafica["Nome"].tolist()
            nome_da_el = st.selectbox("Scegli chi eliminare:", ["Nessuno..."] + lista_nomi)
            if st.button("Elimina", type="primary", width="stretch"):
                if nome_da_el != "Nessuno...":
                    nuovo_df = st.session_state.df_anagrafica[st.session_state.df_anagrafica["Nome"] != nome_da_el]
                    salva_anagrafica(nuovo_df)
                    st.success(f"❌ {nome_da_el} rimosso.")
                    st.rerun()


# ══════════════════════════════════════════════
# SCHEDA — PARAMETRI
# ══════════════════════════════════════════════
with tab_parametri:
    st.subheader("⚙️ Parametri Operativi")

    nuovo_pieces = st.number_input(
        "Pezzi/Ora (Default):",
        value=st.session_state.pieces_ora, min_value=1, key="pieces_ora_input"
    )
    if nuovo_pieces != st.session_state.pieces_ora:
        st.session_state.pieces_ora = nuovo_pieces
        st.rerun()

    st.divider()
    st.subheader("🎯 Forza Lavoro Richiesta (%)")
    st.caption("Percentuale di dipendenti attivi per ogni giorno. Le modifiche richiedono conferma esplicita.")

    sliders_tmp = {}
    col_a, col_b = st.columns(2)
    cols_cycle = [col_a, col_b]
    for idx_g, chiave in enumerate(GIORNI_CHIAVI):
        label = {"Dom_P": "Domenica (inizio)", "Dom_S": "Domenica (fine)"}.get(chiave, chiave)
        valore_corrente = int(round(st.session_state.target_pct_confermato[chiave] * 100))
        with cols_cycle[idx_g % 2]:
            sliders_tmp[chiave] = st.slider(label, 0, 100, valore_corrente, key=f"sl_{chiave}")

    modifiche_pendenti = any(
        sliders_tmp[k] != int(round(st.session_state.target_pct_confermato[k] * 100))
        for k in GIORNI_CHIAVI
    )

    if modifiche_pendenti:
        st.warning("⚠️ Modifiche non applicate")
        col_app, col_ann = st.columns(2)
        with col_app:
            if st.button("✅ Applica modifiche percentuali", type="primary", width="stretch"):
                st.session_state.target_pct_confermato = {k: v / 100 for k, v in sliders_tmp.items()}
                st.rerun()
        with col_ann:
            if st.button("↩️ Annulla modifiche", width="stretch"):
                st.rerun()
